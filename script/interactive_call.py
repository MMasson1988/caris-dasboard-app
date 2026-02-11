#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
CallApp Analysis Script - Version Interactive
Programme d'analyse des appels et visites CARIS avec saisie interactive des dates
"""

import sys
import os
import warnings
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Fill, PatternFill, Protection, Side
from dateutil.relativedelta import relativedelta
from dateutil.parser import parse

warnings.filterwarnings('ignore')

def main():
    print("=== Analyse interactive des appels CARIS ===")
    start_date = input("Date de début (YYYY-MM-DD): ")
    end_date = input("Date de fin (YYYY-MM-DD): ")
    try:
        analyzer = CallAppAnalyzer(start_date, end_date)
        # Ajoutez ici l'appel à la méthode principale d'analyse
        # Ex : analyzer.run_analysis()
        print("Analyse terminée.")
    except Exception as e:
        print(f"Erreur lors de l'analyse: {e}")

if __name__ == '__main__':
    main()

class CallAppAnalyzer:
    """Classe principale pour l'analyse des données CallApp"""
    
    def __init__(self, start_date, end_date):
        """Initialise l'analyseur avec les dates fournies"""
        self.data_folder = Path("./data")
        self.output_folder = Path("./outputs")
        self.output_folder.mkdir(exist_ok=True)
        
        # Dates d'analyse
        self.start_date = start_date
        self.end_date = end_date
        # Utilise la date du jour pour today_date (format YYYY-MM-DD)
        self.today_date = datetime.today().date().strftime('%Y-%m-%d')
        
        # Ordre des mois
        self.mois_ordre = ['January', 'February', 'March', 'April', 'May', 'June', 
                           'July', 'August', 'September', 'October', 'November', 'December']
        
        print("\n" + "="*60)
        print("📊 CARIS Call & Visit Analysis Program")
        print("="*60)
        print(f"📅 Période d'analyse: {self.start_date.date()} → {self.end_date.date()}")
        print(f"📁 Dossier de données: {self.data_folder}")
        print(f"📁 Dossier de sortie: {self.output_folder}")
        print("="*60 + "\n")
    
    @staticmethod
    def get_date_range_from_user():
        """Demande à l'utilisateur de saisir la période d'analyse"""
        print("\n" + "="*60)
        print("🗓️  CONFIGURATION DE LA PÉRIODE D'ANALYSE")
        print("="*60)
        
        while True:
            try:
                print("\n📌 Entrez la date de DÉBUT de l'analyse")
                start_str = input("   Format (AAAA-MM-JJ) ex: 2026-01-12: ").strip()
                start_date = pd.to_datetime(start_str)
                
                print("\n📌 Entrez la date de FIN de l'analyse")
                end_str = input("   Format (AAAA-MM-JJ) ex: 2026-01-18: ").strip()
                end_date = pd.to_datetime(end_str)
                
                # Validation
                if start_date > end_date:
                    print("\n❌ ERREUR: La date de début doit être antérieure à la date de fin!")
                    continue
                
                if end_date > pd.to_datetime(datetime.now()):
                    print("\n⚠️  ATTENTION: La date de fin est dans le futur!")
                    confirm = input("   Voulez-vous continuer? (o/n): ").strip().lower()
                    if confirm != 'o':
                        continue
                
                # Confirmation
                print(f"\n✅ Période sélectionnée: {start_date.date()} → {end_date.date()}")
                print(f"   Durée: {(end_date - start_date).days + 1} jours")
                confirm = input("\n   Confirmer cette période? (o/n): ").strip().lower()
                
                if confirm == 'o':
                    return start_date, end_date
                else:
                    print("\n🔄 Recommençons...\n")
                    
            except Exception as e:
                print(f"\n❌ Erreur de saisie: {e}")
                print("   Veuillez utiliser le format AAAA-MM-JJ (ex: 2026-01-12)\n")
    
    def assign_commune(self, name):
        """Assigne les communes basées sur le préfixe du nom d'utilisateur"""
        import re
        name_str = str(name)
        if re.match(r'^1', name_str):
            return 'Cap-Haïtien'
        elif re.match(r'^2', name_str):
            return 'Port-au-Prince'
        elif re.match(r'^5', name_str):
            return 'Port-de-Paix'
        elif re.match(r'^6', name_str):
            return 'Gonaïves'
        else:
            return 'Autre'
    
    def transform_to_month_year_french(self, df, date_column):
        """Transforme les dates en mois et année en français"""
        if date_column not in df.columns:
            raise ValueError(f"La colonne '{date_column}' n'existe pas dans le DataFrame.")
        df[date_column] = pd.to_datetime(df[date_column])
        df['mois'] = df[date_column].dt.strftime('%B').str.capitalize()
        df['annee'] = df[date_column].dt.strftime('%Y').str.capitalize()
        return df
    
    def dataframe_for_period(self, df, date_column):
        """Filtre le DataFrame par période et assigne les communes"""
        df[date_column] = pd.to_datetime(df[date_column])
        df = df[(df[date_column] >= self.start_date) & (df[date_column] <= self.end_date)]
        df['commune'] = df['username'].apply(self.assign_commune)
        return df
    
    def create_pivot_table(self, df, oev_patient_code):
        """Création du tableau croisé dynamique"""
        pivot_table = pd.pivot_table(
            df,
            values=oev_patient_code,
            index=['mois'],
            columns=['commune', 'Trouvé'],
            aggfunc=lambda x: len(x),
            margins=True,
            margins_name='Total'
        ).fillna(0)
        
        pivot_table = pivot_table.reindex(self.mois_ordre).fillna(0)
        filtered_pivot_table = pivot_table[(pivot_table != 0).any(axis=1)]
        return filtered_pivot_table
    
    def add_found_percentage(self, data_pivotable):
        """Ajoute le pourcentage de personnes trouvées"""
        if 'Total' not in data_pivotable.columns:
            print("⚠️  La colonne 'Total' est manquante dans le DataFrame.")
            return data_pivotable
        
        data_pivotable['Total_Trouvé'] = data_pivotable.xs('Oui', level=1, axis=1).sum(axis=1)
        data_pivotable['% Trouvé'] = data_pivotable['Total_Trouvé'] / data_pivotable['Total']
        data_pivotable['% Trouvé'] = data_pivotable['% Trouvé'].apply(
            lambda x: f"{x:.2%}" if pd.notnull(x) else "0.00%"
        )
        return data_pivotable
    
    def load_datasets(self):
        """Charge tous les fichiers de données avec today_date et chemins absolus"""
        print("📂 Chargement des fichiers de données...\n")
        datasets = {}
        base_path = "C:/Users/Moise/Downloads/caris-dashboard-app/data/"
        today_date = self.today_date
        file_mapping = {
            'Apel_ptme': f"Caris Health Agent - Femme PMTE  - APPELS PTME (created 2025-02-13) {today_date}.xlsx",
            'Apel_oev': f"Caris Health Agent - Enfant - APPELS OEV (created 2025-01-08) {today_date}.xlsx",
            'Visite_ptme': f"Caris Health Agent - Femme PMTE  - Visite PTME (created 2025-02-13) {today_date}.xlsx",
            'Ration_ptme': f"Caris Health Agent - Femme PMTE  - Ration & Autres Visites (created 2025-02-18) {today_date}.xlsx",
            'Ration_oev': f"Caris Health Agent - Enfant - Ration et autres visites (created 2022-08-29) {today_date}.xlsx",
            'oev_visite': f"Caris Health Agent - Enfant - Visite Enfant (created 2025-07-30) {today_date}.xlsx"
        }
        for key, filename in file_mapping.items():
            filepath = base_path + filename
            try:
                print(f"   ⏳ Chargement: {filename[:60]}...")
                datasets[key] = pd.read_excel(filepath, parse_dates=True)
                print(f"   ✓ {key}: {datasets[key].shape[0]} lignes chargées")
            except FileNotFoundError:
                print(f"   ⚠️  FICHIER NON TROUVÉ: {filename}")
                raise
        print("\n✅ Tous les fichiers chargés avec succès!\n")
        return datasets
    
    def process_data(self, datasets):
        """Traite et nettoie les données"""
        print("🔄 Traitement des données...\n")
        
        # Mapping des colonnes pour chaque type de données
        column_mappings = {
            'Apel_ptme': {
                'columns': ['formid','form.APPELS_PTME.patient_code','form.APPELS_PTME.date_appel',
                           'form.APPELS_PTME.is_ptme_available','username','Programme','Type'],
                'rename': {
                    'form.APPELS_PTME.patient_code': 'patient_code',
                    'form.APPELS_PTME.date_appel': 'date',
                    'form.APPELS_PTME.is_ptme_available': 'Trouvé'
                },
                'programme': 'PTME',
                'type': 'Appel'
            },
            'Visite_ptme': {
                'columns': ['formid','form.visite_ptme.health_id','form.visite_ptme.date_of_visit',
                           'form.visite_ptme.is_present','username','Programme','Type'],
                'rename': {
                    'form.visite_ptme.health_id': 'patient_code',
                    'form.visite_ptme.date_of_visit': 'date',
                    'form.visite_ptme.is_present': 'Trouvé'
                },
                'programme': 'PTME',
                'type': 'Visite'
            },
            'Ration_ptme': {
                'columns': ['formid','form.visit_ratio_and_others.patient_code',
                           'form.visit_ratio_and_others.date_of_visit',
                           'form.visit_ratio_and_others.is_benficiary_present','username','Programme','Type'],
                'rename': {
                    'form.visit_ratio_and_others.patient_code': 'patient_code',
                    'form.visit_ratio_and_others.date_of_visit': 'date',
                    'form.visit_ratio_and_others.is_benficiary_present': 'Trouvé'
                },
                'programme': 'PTME',
                'type': 'Visite'
            },
            'Apel_oev': {
                'columns': ['formid','form.appels_oev.patient_code','form.appels_oev.date_appel',
                           'form.appels_oev.parenttuteur_trouve','username','Programme','Type'],
                'rename': {
                    'form.appels_oev.patient_code': 'patient_code',
                    'form.appels_oev.date_appel': 'date',
                    'form.appels_oev.parenttuteur_trouve': 'Trouvé'
                },
                'programme': 'OEV',
                'type': 'Appel'
            },
            'Ration_oev': {
                'columns': ['formid','form.visit_ratio_and_others.patient_code',
                           'form.visit_ratio_and_others.date_of_visit',
                           'form.visit_ratio_and_others.is_benficiary_present','username','Programme','Type'],
                'rename': {
                    'form.visit_ratio_and_others.patient_code': 'patient_code',
                    'form.visit_ratio_and_others.date_of_visit': 'date',
                    'form.visit_ratio_and_others.is_benficiary_present': 'Trouvé'
                },
                'programme': 'OEV',
                'type': 'Visite'
            },
            'oev_visite': {
                'columns': ['formid','form.visite_enfant.patient_code','form.visite_enfant.date_of_visit',
                           'form.visite_enfant.is_available_at_time_visit','username','Programme','Type'],
                'rename': {
                    'form.visite_enfant.patient_code': 'patient_code',
                    'form.visite_enfant.date_of_visit': 'date',
                    'form.visite_enfant.is_available_at_time_visit': 'Trouvé'
                },
                'programme': 'OEV',
                'type': 'Visite'
            }
        }
        
        # Traitement de chaque dataset
        processed_dfs = []
        for key, mapping in column_mappings.items():
            df = datasets[key].copy()
            df['Programme'] = mapping['programme']
            df['Type'] = mapping['type']
            df = df[mapping['columns']]
            df.rename(columns=mapping['rename'], inplace=True)
            processed_dfs.append(df)
            print(f"   ✓ {key}: {df.shape[0]} lignes traitées")
        
        # Combinaison des données
        data = pd.concat(processed_dfs, axis=0, ignore_index=True)
        print(f"\n   ✓ Total combiné: {data.shape[0]} lignes")
        
        # Nettoyage des données
        print("\n🧹 Nettoyage des données...")
        data['date'].replace('---', '1901-01-01', inplace=True)
        data = self.dataframe_for_period(data, 'date')
        print(f"   ✓ Après filtrage par période: {data.shape[0]} lignes")
        
        data = self.transform_to_month_year_french(data, 'date')
        
        # Normalisation des valeurs
        data.replace({'Trouvé': {0: "Non", 1: "Oui", '0': "Non", '1': "Oui"}}, inplace=True)
        
        # Filtrage des données incorrectes
        data_nocorrect = data[data['Trouvé'] == '---']
        data_cleaned = data[(data['Trouvé'] == 'Oui') | (data['Trouvé'] == 'Non')]
        
        if len(data_nocorrect) > 0:
            print(f"   ⚠️  Données incorrectes: {data_nocorrect.shape[0]} lignes")
        
        # Colonnes finales
        cleaned_columns = ['formid','date','Programme','Trouvé','Type','commune', 'mois','annee','username','patient_code']
        data_cleaned = data_cleaned[cleaned_columns]
        
        print("✅ Traitement des données terminé!\n")
        return data_cleaned
    
    def generate_statistics(self, data_cleaned):
        """Génère les statistiques"""
        print("📊 Génération des statistiques...\n")
        stats = {}
        
        # Performances globales
        grouped_data = data_cleaned.groupby(['Programme', 'Type']).size().reset_index(name='Performances')
        print("   Performances par programme et type:")
        for _, row in grouped_data.iterrows():
            print(f"      - {row['Programme']} ({row['Type']}): {row['Performances']}")
        
        # Visites par agent
        agent_data_visite = (
            data_cleaned[data_cleaned['Type'] == "Visite"]
            .groupby(['username'])
            .size()
            .reset_index(name='Performances')
        )
        total_visits = agent_data_visite['Performances'].sum()
        print(f"\n   📍 Total des visites: {total_visits}")
        print(f"   👥 Nombre d'agents: {len(agent_data_visite)}")
        stats['agent_data_visite'] = agent_data_visite
        
        # Appels par agent
        agent_data_appel = (
            data_cleaned[data_cleaned['Type'] == "Appel"]
            .groupby(['username'])
            .size()
            .reset_index(name='Performances')
        )
        total_calls = agent_data_appel['Performances'].sum()
        print(f"\n   📞 Total des appels: {total_calls}")
        print(f"   👥 Nombre d'agents: {len(agent_data_appel)}")
        stats['agent_data_appel'] = agent_data_appel
        
        # Performances totales
        agent_data = (
            data_cleaned.groupby(['username'])
            .size()
            .reset_index(name='Performances')
        )
        stats['agent_data'] = agent_data
        
        print("\n✅ Statistiques générées!\n")
        return stats
    
    def create_pivot_tables(self, data_cleaned):
        """Crée tous les tableaux croisés"""
        print("📋 Création des tableaux croisés...\n")
        pivot_tables = {}
        
        # Tableau total
        print("   ⏳ Tableau total (Appels + Visites)...")
        total_table = data_cleaned[(data_cleaned['Type'] == "Appel") | (data_cleaned['Type'] == "Visite")]
        if not total_table.empty:
            pivot_tables['total'] = self.create_pivot_table(total_table, 'patient_code')
            if isinstance(pivot_tables['total'].columns, pd.MultiIndex) and 'Oui' in pivot_tables['total'].columns.get_level_values(1):
                pivot_tables['total'] = self.add_found_percentage(pivot_tables['total'])
            print("   ✓ Tableau total créé")
        
        # Tableau visites
        print("   ⏳ Tableau des visites...")
        visit_table = data_cleaned[data_cleaned['Type'] == "Visite"]
        if not visit_table.empty:
            pivot_tables['visit'] = self.create_pivot_table(visit_table, 'patient_code')
            print("   ✓ Tableau visites créé")
        
        # Tableau appels
        print("   ⏳ Tableau des appels...")
        appel_table = data_cleaned[data_cleaned['Type'] == "Appel"]
        if not appel_table.empty:
            pivot_tables['appel'] = self.create_pivot_table(appel_table, 'patient_code')
            print("   ✓ Tableau appels créé")
        
        # Tableau visites OEV
        print("   ⏳ Tableau des visites OEV...")
        visite_table_oev = data_cleaned[(data_cleaned['Programme'] == "OEV") & (data_cleaned['Type'] == "Visite")]
        if not visite_table_oev.empty:
            pivot_tables['visite_oev'] = self.create_pivot_table(visite_table_oev, 'patient_code')
            if isinstance(pivot_tables['visite_oev'].columns, pd.MultiIndex) and 'Oui' in pivot_tables['visite_oev'].columns.get_level_values(1):
                pivot_tables['visite_oev'] = self.add_found_percentage(pivot_tables['visite_oev'])
            print("   ✓ Tableau visites OEV créé")
        
        # Tableau PTME visites
        print("   ⏳ Tableau des visites PTME...")
        visite_table_ptme = data_cleaned[(data_cleaned['Programme'] == "PTME") & (data_cleaned['Type'] == "Visite")]
        if not visite_table_ptme.empty:
            pivot_tables['visite_ptme'] = self.create_pivot_table(visite_table_ptme, 'patient_code')
            if isinstance(pivot_tables['visite_ptme'].columns, pd.MultiIndex) and 'Oui' in pivot_tables['visite_ptme'].columns.get_level_values(1):
                pivot_tables['visite_ptme'] = self.add_found_percentage(pivot_tables['visite_ptme'])
            print("   ✓ Tableau visites PTME créé")
        
        # Tableau appels PTME
        print("   ⏳ Tableau des appels PTME...")
        appel_table_ptme = data_cleaned[(data_cleaned['Programme'] == "PTME") & (data_cleaned['Type'] == "Appel")]
        if not appel_table_ptme.empty:
            pivot_tables['appel_ptme'] = self.create_pivot_table(appel_table_ptme, 'patient_code')
            if isinstance(pivot_tables['appel_ptme'].columns, pd.MultiIndex) and 'Oui' in pivot_tables['appel_ptme'].columns.get_level_values(1):
                pivot_tables['appel_ptme'] = self.add_found_percentage(pivot_tables['appel_ptme'])
            print("   ✓ Tableau des appels PTME créé")
        
        # Tableau appels OEV
        print("   ⏳ Tableau des appels OEV...")
        appel_table_oev = data_cleaned[(data_cleaned['Programme'] == "OEV") & (data_cleaned['Type'] == "Appel")]
        if not appel_table_oev.empty:
            pivot_tables['appel_oev'] = self.create_pivot_table(appel_table_oev, 'patient_code')
            if isinstance(pivot_tables['appel_oev'].columns, pd.MultiIndex) and 'Oui' in pivot_tables['appel_oev'].columns.get_level_values(1):
                pivot_tables['appel_oev'] = self.add_found_percentage(pivot_tables['appel_oev'])
            print("   ✓ Tableau des appels OEV créé")
        
        print("\n✅ Tous les tableaux croisés créés!\n")
        return pivot_tables
    
    def export_to_excel(self, data_cleaned, stats, pivot_tables):
        """Exporte tous les résultats vers Excel"""
        print("💾 Export des résultats vers Excel...\n")
        
        output_file = self.output_folder / f"Analyse_Appels_Visites_{self.start_date.date()}_to_{self.end_date.date()}.xlsx"
        
        print(f"   📄 Fichier: {output_file.name}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Export des données nettoyées
            data_cleaned.to_excel(writer, sheet_name='Données Nettoyées', index=False)
            
            # Export des statistiques
            grouped_data = data_cleaned.groupby(['Programme', 'Type']).size().reset_index(name='Performances')
            grouped_data.to_excel(writer, sheet_name='Statistiques Globales', index=False)
            
            if 'agent_data_visite' in stats:
                stats['agent_data_visite'].to_excel(writer, sheet_name='Visites par Agent', index=False)
            
            if 'agent_data_appel' in stats:
                stats['agent_data_appel'].to_excel(writer, sheet_name='Appels par Agent', index=False)
            
            if 'agent_data' in stats:
                stats['agent_data'].to_excel(writer, sheet_name='Performance Totale', index=False)
            
            print("   ✅ Statistiques exportées")
            
            # Export des tableaux croisés
            print("   ⏳ Export des tableaux croisés...")
            if 'total' in pivot_tables:
                pivot_tables['total'].to_excel(writer, sheet_name='Tableau Total')
            
            if 'visit' in pivot_tables:
                pivot_tables['visit'].to_excel(writer, sheet_name='Tableau Visites')
            
            if 'appel' in pivot_tables:
                pivot_tables['appel'].to_excel(writer, sheet_name='Tableau Appels')
            
            if 'visite_ptme' in pivot_tables:
                pivot_tables['visite_ptme'].to_excel(writer, sheet_name='Visites PTME')
            
            if 'appel_ptme' in pivot_tables:
                pivot_tables['appel_ptme'].to_excel(writer, sheet_name='Appels PTME')
            
            if 'visite_oev' in pivot_tables:
                pivot_tables['visite_oev'].to_excel(writer, sheet_name='Visites OEV')
            
            if 'appel_oev' in pivot_tables:
                pivot_tables['appel_oev'].to_excel(writer, sheet_name='Appels OEV')
            
            print("   ✅ Tableaux croisés exportés")
        
        print(f"\n✅ Fichier Excel créé avec succès!")
        print(f"📁 Emplacement: {output_file.absolute()}\n")
        
        return output_file
    
    def run(self):
        """Exécute l'analyse complète"""
        try:
            # Chargement des données
            datasets = self.load_datasets()
            
            # Traitement des données
            data_cleaned = self.process_data(datasets)
            
            # Génération des statistiques
            stats = self.generate_statistics(data_cleaned)
            
            # Création des tableaux croisés
            pivot_tables = self.create_pivot_tables(data_cleaned)
            
            # Export vers Excel
            output_file = self.export_to_excel(data_cleaned, stats, pivot_tables)
            
            # Résumé final
            print("="*60)
            print("🎉 ANALYSE TERMINÉE AVEC SUCCÈS!")
            print("="*60)
            print(f"📊 Résumé de l'analyse:")
            print(f"   • Période: {self.start_date.date()} → {self.end_date.date()}")
            print(f"   • Total de lignes traitées: {len(data_cleaned)}")
            print(f"   • Fichier de sortie: {output_file.name}")
            print("="*60)
            
            return True
            
        except FileNotFoundError as e:
            print(f"\n❌ ERREUR: Fichier non trouvé - {str(e)}")
            print("   Veuillez vérifier que tous les fichiers nécessaires sont dans le dossier 'data'")
            return False
            
        except Exception as e:
            print(f"\n❌ ERREUR INATTENDUE: {str(e)}")
            import traceback
            print("\n📝 Détails de l'erreur:")
            traceback.print_exc()
            return False


def main():
    """Point d'entrée principal du programme"""
    print("\n" + "🚀 "*20)
    print("DÉMARRAGE DU PROGRAMME D'ANALYSE CALLAPP")
    print("🚀 "*20 + "\n")
    
    # Demander les dates AVANT de créer l'analyseur
    start_date, end_date = CallAppAnalyzer.get_date_range_from_user()
    
    # Créer et exécuter l'analyseur avec les dates fournies
    analyzer = CallAppAnalyzer(start_date, end_date)
    success = analyzer.run()
    
    if success:
        print("\n✅ Programme terminé avec succès!")
    else:
        print("\n⚠️  Programme terminé avec des erreurs.")
    
    input("\n👉 Appuyez sur Entrée pour quitter...")


if __name__ == "__main__":
    main()
