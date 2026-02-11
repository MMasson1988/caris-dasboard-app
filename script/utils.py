from difflib import SequenceMatcher
from typing import List
import requests
import json
import pandas as pd
from dotenv import load_dotenv
import os
from datetime import datetime, timedelta
# Get system timezone
import time
import re
load_dotenv()
from sqlalchemy import create_engine
from pathlib import Path
#====================================================================================================
# Obtenir la date/heure actuelle
now = datetime.now()

# Si l'heure est 24h (minuit) ou plus → on ajoute un jour
if now.hour >= 24:
    now += timedelta(days=1)

# Formater la date finale
today_str = now.strftime("%Y-%m-%d")
#====================================================================================================
def write_excel(df, output_file):
    output_path = Path(output_file)
    if not output_path.is_absolute():
        output_path = Path(__file__).resolve().parent.parent / output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"✅ Fichier enregistré : {output_path}")
#====================================================================================================

def load_excel_to_df(filename: str, df_name: str) -> pd.DataFrame:
    """
    Charge un fichier Excel situé à la racine du projet (ou dans un sous-dossier)
    et renvoie le DataFrame correspondant.
    
    Args:
        filename (str): nom du fichier à lire (ex: 'nutrition.xlsx' ou 'data/nutrition.xlsx')
        df_name (str): nom symbolique du DataFrame à afficher dans les logs
    
    Returns:
        pd.DataFrame: le DataFrame chargé depuis le fichier Excel
    """
    # Déterminer la racine du projet (racine du dossier)
    base_dir = Path(__file__).resolve().parent.parent
    data_path = base_dir / "data" / filename
    if not data_path.exists():
        raise FileNotFoundError(f"❌ Fichier introuvable : {data_path}")
    print(f"📂 Lecture du fichier '{filename}' dans {data_path}")
    df = pd.read_excel(data_path)
    print(f"✅ DataFrame '{df_name}' chargé ({len(df)} lignes, {len(df.columns)} colonnes)")
    return df

if __name__ == '__main__':
    print("Module utils.py chargé. Ajoutez une fonction main() pour exécuter des tests ou des exemples.")
#=====================================================================================================
def get_commcare_odata(url, auth_credentials, filter_params):
    """
    Fetch active muso groups from CommCare using OData API
    
    Args:
        url (str): The OData API URL
        auth_credentials (tuple): Username and password tuple (username, password)
        filter_params (dict): Parameters to filter the data
        
    Returns:
        list: List of muso group records
    """
    # Make the initial request to the OData API
    response = requests.get(url, auth=auth_credentials, params=filter_params)

    # Check if the request was successful
    if response.status_code != 200:
        print(f"Error: API request failed with status code {response.status_code}")
        return []
    # Get initial data
    data = response.json()['value']
    # Follow pagination links if they exist
    while "@odata.nextLink" in response.json():
        next_link = response.json()["@odata.nextLink"]
        print(f"Following next link: {next_link}")
        
        # Get the next page of data
        response = requests.get(next_link, auth=auth_credentials)
        
        if response.status_code == 200:
            # Add new records to our data
            new_records = response.json()['value']
            data += new_records
            print(f"Retrieved additional {len(new_records)} records. Total: {len(data)}")
        else:
            print(f"Error: Failed to retrieve next page: {response.status_code}")
            break
    
    print(f"Total records retrieved: {len(data)}")
    return data

#=========================================================================================================

def is_beneficiary_active(row):
    """
    Check if a beneficiary is active based on various date fields
    Args:
        row (dict): A dictionary representing a beneficiary record
        start_date (str): The start date for the active period
        end_date (str): The end date for the active period"
        """
    if row['closed_date']<start_date:
        return "no"
    if row["creation_date"]>end_date:
        return "no"
    if row['graduation_date']>start_date:
        return "yes"
    if row['abandoned_date']>start_date:
        return "yes"
    if row['inactive_date']>start_date:
        return "yes"
    if row['inactive_date']<start_date:
        return "no"
    if row['graduation_date']<start_date:
        return "no"
    if (pd.isnull(row['is_inactive']) or row['is_inactive']==0) and (row['graduated']==0 or pd.isnull(row['graduated'])):
        return "yes"
    return "no"


# define a function to check if a group is active
def is_groupe_active(row):
    """
    Check if a group is active based on various date fields
    Args:
        row (dict): A dictionary representing a group record
    Returns:
        str: 'yes' if active, 'no' if inactive
    """
        # Convert reference dates to UTC timezone-aware timestamps
    start_date_dt = pd.to_datetime(start_date)
    end_date_dt = pd.to_datetime(end_date)
    print("start date and end date", start_date_dt, end_date_dt)
    
    # Convert and normalize dates from the row
    # Check conditions only if dates are valid
    if row["office_name"] in ["CAY","JER"]:
        return "no"
    # if pd.notna(row['opened_date']) and row['opened_date'] < start_date_dt:
    #     return "no"
    if pd.notna(row['closed_date']) and row['closed_date'] < start_date_dt:
        return "no"
    if pd.notna(row['creation_date']) and row["creation_date"] > end_date_dt:
        return "no"
    if pd.notna(row['graduation_date']) and row['graduation_date'] > start_date_dt:
        return "yes"
    if pd.notna(row['inactive_date']) and row['inactive_date'] > start_date_dt:
        return "yes"
    if pd.notna(row['inactive_date']) and row['inactive_date'] < start_date_dt:
        return "no"
    if pd.notna(row['graduation_date']) and row['graduation_date'] < start_date_dt:
        return "no"
    if (pd.isnull(row['is_inactive']) or row['is_inactive']==0) and (row['is_graduated']==0 or pd.isnull(row['is_graduated'])):
        return "yes"
    return "no"
#=============================================================================================================
def is_screened_in_period(df, date_col, start_date, end_date, ref_date):
    """
    Filter records based on screening period logic
    """
    # Convert dates to timestamps
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)
    ref_date = pd.to_datetime(ref_date)
    
    # Ensure date column is datetime
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    
    # Create boolean mask for the condition
    condition = (
        pd.notna(df[date_col]) &
        (df[date_col] >= start_date) &
        (df[date_col] <= end_date) &
        (df[date_col] >= ref_date)
    )
    
    # Apply the filter
    filtered_df = df[condition]
    print(f"Records in screening period: {len(filtered_df)}/{len(df)}")
    
    return filtered_df

#===========================================================================================================================

def check_files(files):
    path = "C:\\Users\\Moise\\Downloads\\Reports_MEAL\\OEV"
    missing_files = []

    for file in files:
        full_path = os.path.join(path, file)
        print(f"Checking for file: {full_path}")  # Debugging print
        if not os.path.isfile(full_path):
            missing_files.append(file)

    if not missing_files:
        print("All files are present.")
    else:
        print(f"The following files are missing: {missing_files}")

# Get the current date in the format YYYY-MM-DD
today_date = datetime.today().strftime('%Y-%m-%d')
#==================================================================================================================
import pandas as pd
def save_dataframe_to_excel(df, filename, output_path="../outputs"):
    # Vérifier si le nom du fichier a l'extension .xlsx, sinon l'ajouter
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    # Construire le chemin complet du fichier
    output_file_path = os.path.join(output_path, filename)

    # Sauvegarder le DataFrame dans un fichier Excel
    df.to_excel(output_file_path, index=False)

    print(f"DataFrame saved to {output_file_path}")
    return output_file_path

# Exemple d'utilisation :
#save_dataframe_to_excel(stat_ptme_index, "stat_ptme_index.xlsx")
#==================================================================================================================
import pandas as pd

def creer_colonne_match_conditional(df1, df2, on, nouvelle_colonne, mapping):
    """
    Crée une colonne dans df1 en fonction des correspondances avec df2, conditionnées par un mapping.
    
    :param df1: Le premier DataFrame
    :param df2: Le second DataFrame
    :param on: La colonne utilisée pour le merge
    :param nouvelle_colonne: Le nom de la nouvelle colonne à créer
    :param mapping: Un dictionnaire pour définir les valeurs de la colonne 'match'
    :return: df1 avec la nouvelle colonne ajoutée
    """
    # Merge gauche avec indicateur de correspondance
    merged_df = df1.merge(df2, on=on, how='left', indicator=True)
    
    # Création de la nouvelle colonne 'match' avec un mapping conditionnel
    merged_df[nouvelle_colonne] = merged_df['_merge'].map(mapping)
    
    # Suppression de la colonne '_merge'
    merged_df.drop(columns=['_merge'], inplace=True)
    
    # Retourner seulement les colonnes originales de df1 avec la nouvelle colonne
    return merged_df

#========================================Functions=========================================================
import plotly.express as px
import plotly.graph_objects as go

def plot_txcurr_by_office(df, title='🎯 TX_CURR: Number of Patients by Office'):
    """
    Affiche un graphique en barres horizontales du nombre de patients par bureau (office).
    
    Paramètres :
    - df : DataFrame contenant au moins les colonnes 'office' et 'patient_code'
    - title : Titre du graphique (optionnel)
    """
    # Regrouper et trier
    txcurr_by_office = df.groupby('office')['patient_code'].count().sort_values(ascending=True)

    # Couleurs
    colors = px.colors.qualitative.Plotly
    color_list = [colors[i % len(colors)] for i in range(len(txcurr_by_office))]

    # Création du graphique
    fig = go.Figure(go.Bar(
        x=txcurr_by_office.values,
        y=txcurr_by_office.index,
        orientation='h',
        marker=dict(
            color=color_list,
            line=dict(color='rgba(0,0,0,0.8)', width=1)
        ),
        text=txcurr_by_office.values,
        textposition='auto',
        textfont=dict(size=14, color='white'),
        hovertemplate='<b>%{y}</b><br>Patients: %{x}<extra></extra>'
    ))

    # Mise en page
    fig.update_layout(
        title=dict(
            text=title,
            x=0.5,
            xanchor='center',
            font=dict(size=22)
        ),
        xaxis=dict(
            title='Number of Patients',
            showgrid=False,
            zeroline=False
        ),
        yaxis=dict(
            title='Office',
            tickfont=dict(size=14, color='black'),
            showgrid=False,
            zeroline=False
        ),
        template='plotly_white',
        height=600,
        width=950,
        margin=dict(l=100, r=40, t=60, b=40)
    )

    fig.show()
    
#========================================Functions=========================================================
import pandas as pd
import plotly.graph_objects as go

def plot_viral_load_summary2_plotly(df, age_min=0, age_max=19, start_date=None, end_date=None, title=None, output='plot'):
    df = df.copy()
    df['age'] = pd.to_numeric(df['age'], errors='coerce')
    df['last_viral_load_collection_date'] = pd.to_datetime(df['last_viral_load_collection_date'], errors='coerce')
    df['arv_start_date'] = pd.to_datetime(df['arv_start_date'], errors='coerce')

    # Filtrage par âge
    df = df[(df['age'] >= age_min) & (df['age'] <= age_max)]
    if df.empty:
        print(f"Aucune donnée pour les patients âgés de {age_min} à {age_max} ans.")
        return None

    # Dates
    end_date = pd.to_datetime('today')

    # Sous-ensembles
    on_arv = df[df['arv_start_date'] <= end_date - pd.DateOffset(months=2)]
    not_on_arv = df[~df.index.isin(on_arv.index)]

    vl_coverage = on_arv[on_arv['last_viral_load_collection_date'] >= end_date - pd.DateOffset(months=12)]
    not_covered = on_arv[~on_arv.index.isin(vl_coverage.index)]

    suppression_col = 'indetectable_ou_inf_1000'
    if df[suppression_col].dtype != 'O':
        df[suppression_col] = df[suppression_col].astype(str)

    vl_suppression = vl_coverage[vl_coverage[suppression_col].str.upper() == "OUI"]
    not_suppressed = vl_coverage[~vl_coverage.index.isin(vl_suppression.index)]

    # Gestion des sorties alternatives
    if output == 'on_arv':
        return on_arv
    elif output == 'not_on_arv':
        return not_on_arv
    elif output == 'vl_coverage':
        return vl_coverage
    elif output == 'not_covered':
        return not_covered
    elif output == 'vl_suppression':
        return vl_suppression
    elif output == 'not_suppressed':
        return not_suppressed
    elif output == 'TX_CURR':
        return df

    # Données pour le graphe
    Indicators = ['TX_CURR', 'On ARV ≥3mois', 'Viral Load Coverage', 'Viral Load Suppression']
    n_1 = [df.shape[0], on_arv.shape[0], vl_coverage.shape[0], vl_suppression.shape[0]]
    prop_impact1 = [
        1.0,
        on_arv.shape[0] / df.shape[0] if df.shape[0] > 0 else 0,
        vl_coverage.shape[0] / on_arv.shape[0] if on_arv.shape[0] > 0 else 0,
        vl_suppression.shape[0] / vl_coverage.shape[0] if vl_coverage.shape[0] > 0 else 0
    ]

    med = pd.DataFrame({
        'Indicators': Indicators,
        'n_1': n_1,
        'prop_impact1': [round(p * 100, 1) for p in prop_impact1]
    })

    # Couleurs personnalisées cohérentes
    color_map = {
        'TX_CURR': "#77198C",
        'On ARV ≥3mois': '#EF553B',
        'Viral Load Coverage': "#00C2CC",
        'Viral Load Suppression': "#F7FA63"
    }
    med['Color'] = med['Indicators'].map(color_map)
    med['Label'] = med.apply(lambda row: f"{row['n_1']:,} ({row['prop_impact1']}%)", axis=1)

    # Titre dynamique ou personnalisé
    if title is None:
        title = f"Cascade des OEV âgés de {age_min}-{age_max} ans"

    # Tracé Plotly
    fig = go.Figure()
    for _, row in med.iterrows():
        fig.add_trace(go.Bar(
            x=[row['Indicators']],
            y=[row['n_1']],
            name=row['Indicators'],
            marker_color=row['Color'],
            text=row['Label'],
            textposition='outside'
        ))

    fig.update_layout(
        title=title,
        xaxis_title="",
        yaxis_title="Fréquence",
        title_x=0.5,
        showlegend=False,
        template='plotly_white',
        margin=dict(t=60, b=60),
        annotations=[
            dict(
                text=f"Source: hivhaiti / {end_date.date()}",
                x=0.5,
                y=-0.2,
                showarrow=False,
                xref='paper',
                yref='paper',
                font=dict(size=11, color="gray")
            )
        ]
    )

    fig.show()
#=======================================Functions=========================================================
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import numpy as np
import pandas as pd
import numpy as np
import plotly.graph_objects as go

def plot_age_pyramid(df, age_col='age', sex_col='sex', id_col='patient_code',
                     title='👥 TX_CURR - Pyramide des âges par sexe (0-17 ans)'):

    # Nettoyage
    df_clean = df[[age_col, sex_col, id_col]].dropna()
    df_clean[age_col] = pd.to_numeric(df_clean[age_col], errors='coerce')
    df_clean = df_clean[df_clean[age_col].between(0, 17)]  # max 17 ans

    # Bins avec dernier groupe 15-17
    bins = [0, 5, 10, 15, 18]
    labels = ['0-4', '5-9', '10-14', '15-17']
    df_clean['age_group'] = pd.cut(df_clean[age_col], bins=bins, labels=labels, right=False)

    # Comptage
    pyramid = df_clean.groupby(['age_group', sex_col])[id_col].count().unstack(fill_value=0)

    if 'F' not in pyramid.columns: pyramid['F'] = 0
    if 'M' not in pyramid.columns: pyramid['M'] = 0

    pyramid = pyramid.sort_index()

    # Traces Plotly
    fig = go.Figure()

    fig.add_trace(go.Bar(
        y=pyramid.index.astype(str),
        x=-pyramid['M'],
        name='Hommes',
        orientation='h',
        marker_color='steelblue',
        text=pyramid['M'],
        textposition='inside',
        textfont=dict(color='white', size=13),
        width=0.9,
        hovertemplate='Âge: %{y}<br>Hommes: %{text}<extra></extra>'
    ))

    fig.add_trace(go.Bar(
        y=pyramid.index.astype(str),
        x=pyramid['F'],
        name='Femmes',
        orientation='h',
        marker_color='lightcoral',
        text=pyramid['F'],
        textposition='inside',
        textfont=dict(color='white', size=13),
        width=0.9,
        hovertemplate='Âge: %{y}<br>Femmes: %{text}<extra></extra>'
    ))

    # Axes dynamiques
    x_max = max(pyramid['M'].max(), pyramid['F'].max())
    step = int(np.ceil(x_max / 4.0 / 10.0)) * 10
    ticks = list(range(-x_max, x_max + step, step))
    ticktext = [abs(v) for v in ticks]

    fig.update_layout(
        title=dict(
            text=title,
            x=0.5,
            xanchor='center',
            font=dict(size=22)
        ),
        barmode='relative',
        xaxis=dict(
            title='Nombre de patients',
            tickvals=ticks,
            ticktext=ticktext,
            zeroline=True,
            showgrid=False
        ),
        yaxis=dict(
            title='Groupe d’âge',
            tickfont=dict(size=13),
            showgrid=False
        ),
        height=600,
        width=950,
        template='plotly_white',
        margin=dict(l=100, r=40, t=60, b=40)
    )

    fig.show()
#===========================================================================================================================
import pandas as pd

def filter_oev_data(
    df,
    excluded_offices=['BOM', 'PDP'],
    excluded_networks=['PIH', 'UGP', 'MSPP'],
    excluded_sites=['PAP/CHAP', 'PAP/OBCG', 'PAP/OGRE', 'PEG/HNDP', 'PAP/SMFO', 'LEG/HSCL', 'PAP/HAHD', 'ARC/SADA'],
    abandoned_flag=1,
    output_file='TX_CURR.xlsx'
):
    print(f"Initial dataset has {df.shape[0]} observations")

    df = df[~df['office'].isin(excluded_offices)]
    print(f"After excluding offices: {df.shape[0]} observations")

    df = df[~df['network'].isin(excluded_networks)]
    print(f"After excluding networks: {df.shape[0]} observations")

    df['age'] = pd.to_numeric(df['age'], errors='coerce')
    df = df[df['age'].between(0, 17)]
    print(f"After filtering age between 0-17: {df.shape[0]} observations")

    df = df[~df['site'].isin(excluded_sites)]
    print(f"After excluding sites: {df.shape[0]} observations")

    df['is_abandoned'] = pd.to_numeric(df['is_abandoned'], errors='coerce')
    df = df[df['is_abandoned'] != abandoned_flag]
    print(f"After excluding abandoned: {df.shape[0]} observations")

    next_appointment_date_max = pd.to_datetime('today')
    first_day_this_month = next_appointment_date_max.replace(day=1)
    next_appointment_date_min = first_day_this_month - pd.DateOffset(months=1)

    df['next_appointment_date'] = pd.to_datetime(df['next_appointment_date'], errors='coerce')
    df = df[df['next_appointment_date'] >= next_appointment_date_min]
    df = df.drop_duplicates(subset=['patient_code'], keep='last')
    print(f"After filtering by next appointment date ≥ {next_appointment_date_min.date()}: {df.shape[0]} observations")

    df.to_excel(output_file, index=False)
    print(f"Filtered data exported to {output_file}")

    return df

#===========================================================================================
def save_dataframes_excel(output_name="output", df1=None, df2=None, df3=None, df4=None, df5=None, df6=None, sheet_names=None):
    """
    Sauvegarde jusqu'à 6 DataFrames dans un fichier Excel avec un nom dynamique basé sur la date.

    Args:
        output_name (str): Nom de base du fichier de sortie (sans extension).
        df1, df2, df3, df4, df5, df6 (pd.DataFrame, optional): DataFrames à sauvegarder.
        sheet_names (list, optional): Liste des noms de feuilles.
    """
    # Définir une date pour le nom du fichier
    today_str = date.today().strftime("%Y-%m-%d")
    file_name = f"{output_name}_{today_str}.xlsx"

    # Définir les noms des feuilles si non fournis
    sheet_names = sheet_names or ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5', 'Sheet6']

    # Créer l'écrivain Excel
    writer = pd.ExcelWriter(file_name, engine='xlsxwriter')

    # Stocker les DataFrames dans une liste
    dfs = [df1, df2, df3, df4, df5, df6]

    # Écrire chaque DataFrame dans sa feuille correspondante s'il est défini
    for i, df in enumerate(dfs):
        if df is not None:
            df.to_excel(writer, sheet_name=sheet_names[i], index=False)

    # Sauvegarder le fichier Excel
    writer.close()
    print(f"Fichier '{file_name}' sauvegardé avec succès.")
    
#===========================================================================================
import pandas as pd
import unicodedata
from rapidfuzz import process, fuzz

def commcare_match_person(
    df_reference: pd.DataFrame,
    df_commcare: pd.DataFrame,
    name_column: str = "name",
    threshold: int = 85,
    return_df: str = "reference",  # "reference" ou "commcare"
    scorer=fuzz.token_set_ratio
) -> pd.DataFrame:
    """
    Apparier flou sur une colonne de noms entre df_reference et df_commcare.

    Args:
        df_reference: DataFrame source de référence.
        df_commcare: DataFrame cible où l'on cherche la meilleure correspondance.
        name_column: Nom de la colonne contenant les noms à comparer (défaut: 'name').
        threshold: Score minimal [0..100] pour accepter une correspondance.
        return_df: 'reference' => retourne df_reference enrichi;
                   'commcare'     => retourne df_commcare enrichi (recherche inversée).
        scorer: fonction de similarité rapidfuzz (par défaut token_set_ratio).

    Returns:
        Le DataFrame demandé, avec 3 colonnes ajoutées:
          - 'best_match'    : le nom correspondant le plus proche
          - 'score'         : le score de similarité (0..100)
          - 'correspondance': 'yes' si score >= threshold, 'no' sinon
    """

    if "name" not in df_reference.columns or "name" not in df_commcare.columns:
        raise ValueError("Les deux DataFrames doivent contenir la colonne 'name'.")

    def _normalize(s: str) -> str:
        if pd.isna(s):
            return ""
        s = str(s).strip().lower()
        s = " ".join(s.split())  # collapse espaces multiples
        # retirer les accents
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        return s

    def _match(source_df: pd.DataFrame, target_df: pd.DataFrame) -> pd.DataFrame:
        # Prépare les choix (normalisés) et un mapping -> original
        target_norm = target_df[name_column].map(_normalize).fillna("")
        choices = target_norm.tolist()
        orig_by_norm = target_norm.to_frame("norm").join(
            target_df[name_column]
        ).drop_duplicates(subset=["norm"]).set_index("norm")[name_column].to_dict()

        # Calcul du meilleur match pour chaque source
        best_matches = []
        scores = []

        for name in source_df[name_column].fillna(""):
            q = _normalize(name)
            if not q:
                best_matches.append(None)
                scores.append(0)
                continue

            # extractOne renvoie (match_norm, score, index)
            match = process.extractOne(q, choices, scorer=scorer)
            if match is None:
                best_matches.append(None)
                scores.append(0)
                continue

            match_norm, score, _ = match
            if score >= threshold:
                best_matches.append(orig_by_norm.get(match_norm, None))
                scores.append(int(score))
            else:
                best_matches.append(None)
                scores.append(int(score))

        out = source_df.copy()
        out["best_match"] = best_matches
        out["score"] = scores
        out["correspondance"] = ["yes" if score >= threshold else "no" for score in scores]
        return out

    if return_df not in ("reference", "commcare"):
        raise ValueError("return_df doit être 'reference' ou 'commcare'.")

    if return_df == "reference":
        # cherche pour chaque nom du df_reference dans df_commcare
        return _match(df_reference, df_commcare)
    else:
        # cherche pour chaque nom du df_commcare dans df_reference
        return _match(df_commcare, df_reference)
#===========================================================================================================================
def classify_mam_mas_muac(df, muac_col="muac_mm", am_class_col="AM_class"):
    """
    Classe les enfants selon leur MUAC (OMS) :
      - MAS si MUAC < 115 mm
      - MAM si 115 ≤ MUAC < 125 mm
      - Aucune AM si MUAC ≥ 125 mm
      - Non applicable si MUAC manquant

    Args:
        df : DataFrame source
        muac_col : nom de la colonne contenant le PB (MUAC)
        am_class_col : nom de la colonne où stocker le résultat
    """

    # Convertir la colonne MUAC en numérique (au cas où certaines valeurs seraient des chaînes)
    df.loc[:, muac_col] = pd.to_numeric(df[muac_col], errors='coerce')

    # Utiliser np.where pour classer selon les seuils
    df.loc[:, am_class_col] = np.where(df[muac_col].isna(), "MAM",
                        np.where(df[muac_col] < 11.5, "MAS",
                        np.where(df[muac_col] < 12.5, "MAM", "MAM")))

    return df


#=================================================================================================
def extraire_data(df, start_date, end_date, date_col):
    # Conversion des dates d'entrée
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)

    # Conversion de la colonne de date dans le DataFrame
    if date_col in df.columns:
        df.loc[:, date_col] = pd.to_datetime(df[date_col], errors='coerce')
    else:
        raise ValueError(f"Colonne '{date_col}' absente du DataFrame.")

    # Filtrage
    df = df[df[date_col].between(start_date, end_date)]

    return df
#=============================================
import pandas as pd

def clean_column_names(df, expr_to_remove):
    """
    Supprime une expression donnée au début des noms de colonnes d'un DataFrame.

    Args:
        df (pd.DataFrame): le dataframe cible
        expr_to_remove (str): l'expression à supprimer (ex: 'form.depistage.')

    Returns:
        pd.DataFrame: un DataFrame avec colonnes renommées
    """
    df = df.rename(columns=lambda col: col.replace(expr_to_remove, ""))
    return df
#========================================================================================
import pandas as pd

def create_binary_symptom_columns(df, col="symptoms"):
    """
    Transforme une colonne de type string en colonnes indicatrices binaires (0/1).
    - Si 'aucun' est présent avec d'autres modalités => on ignore 'aucun'.
    - Si 'aucun' est la seule modalité => on crée une colonne 'aucun'=1.
    
    Args:
        df (pd.DataFrame): le DataFrame source
        col (str): nom de la colonne contenant les modalités séparées par espaces

    Returns:
        pd.DataFrame: DataFrame avec colonnes indicatrices ajoutées
    """
    # Faire une copie explicite pour éviter SettingWithCopyWarning
    df = df.copy()
    
    # Créer une liste de sets de symptômes par ligne
    df.loc[:, col] = df[col].fillna("").str.lower()
    df.loc[:, col] = df[col].str.strip()
    df.loc[:, col] = df[col].str.split()

    # Extraire toutes les modalités distinctes
    all_modalities = set(mod for mods in df[col] for mod in mods)

    # On retire 'aucun' provisoirement
    if "aucun" in all_modalities:
        all_modalities.remove("aucun")

    # Créer colonnes indicatrices pour chaque modalité (hors 'aucun')
    for mod in all_modalities:
        df.loc[:, mod] = df[col].apply(lambda mods: 1 if mod in mods else 0)

    # Gérer la règle spéciale pour 'aucun'
    df.loc[:, "aucun"] = df[col].apply(lambda mods: 1 if mods == ["aucun"] else 0)

    return df
#============================================================================
def get_age_in_year(df, dob_col):
    """
    Calcule l'âge en années pour chaque ligne d'un DataFrame où la colonne date de naissance n'est pas NaN.
    
    Args:
        df (pd.DataFrame): le DataFrame d'entrée
        dob_col (str): nom de la colonne contenant la date de naissance
        
    Returns:
        pd.DataFrame: DataFrame avec une nouvelle colonne 'age_years'
    """
    # Faire une copie du DataFrame pour éviter les modifications sur l'original
    df_copy = df.copy()
    
    # Vérifier que la colonne existe
    if dob_col not in df_copy.columns:
        raise ValueError(f"Colonne '{dob_col}' absente du DataFrame.")
    
    # Convertir la colonne date de naissance en datetime
    df_copy[dob_col] = pd.to_datetime(df_copy[dob_col], errors='coerce')
    
    # Calculer l'âge seulement là où dob_col n'est pas NaN
    today = pd.Timestamp.now()
    
    # Créer la colonne age_years avec des 0 par défaut (entier)
    df_copy['age_years'] = 0
    
    # Calculer l'âge seulement pour les lignes où date de naissance n'est pas NaN
    mask = df_copy[dob_col].notna()
    age_values = ((today - df_copy.loc[mask, dob_col]).dt.days / 365.25).round(0).astype(int)
    df_copy.loc[mask, 'age_years'] = age_values
    return df_copy

#=====================================================================
def get_age_in_months(df, dob_col):
    """
    Calcule l'âge en mois pour chaque ligne d'un DataFrame où la colonne date de naissance n'est pas NaN.
    
    Args:
        df (pd.DataFrame): le DataFrame d'entrée
        dob_col (str): nom de la colonne contenant la date de naissance
        
    Returns:
        pd.DataFrame: DataFrame avec une nouvelle colonne 'age_months'
    """
    # Faire une copie du DataFrame pour éviter les modifications sur l'original
    df_copy = df.copy()
    
    # Vérifier que la colonne existe
    if dob_col not in df_copy.columns:
        raise ValueError(f"Colonne '{dob_col}' absente du DataFrame.")
    
    # Convertir la colonne date de naissance en datetime
    df_copy[dob_col] = pd.to_datetime(df_copy[dob_col], errors='coerce')
    
    # Calculer l'âge seulement là où dob_col n'est pas NaN
    today = pd.Timestamp.now()
    
    # Créer la colonne age_months avec des 0 par défaut (entier)
    df_copy['age_months'] = 0
    
    # Calculer l'âge en mois seulement pour les lignes où date de naissance n'est pas NaN
    mask = df_copy[dob_col].notna()
    age_months_values = ((today - df_copy.loc[mask, dob_col]).dt.days / 30.44).round(0).astype(int)
    df_copy.loc[mask, 'age_months'] = age_months_values
    
    return df_copy
#=====================================================================
import pandas as pd

def create_normalized_health_index(df):
    """
    Crée un indice pondéré et normalisé (0-10) basé sur symptômes et variables nutritionnelles,
    puis ajoute une catégorisation de risque.
    """
    # Faire une copie explicite pour éviter SettingWithCopyWarning
    df = df.copy()

    # Pondération malnutrition
    df["malnutrition_score"] = df["manutrition_type"].map(
        {"MAM": 2, "MAS": 4}
    ).fillna(0)

    # Pondération diarrhée
    df["diarrhea_score"] = df["diarrhea"].map(
        {"yes": 3, "no": 0}
    ).fillna(0)

    # Pondération oedème
    df["oedem_score"] = df["edema"].map(
        {"yes": 5, "no": 0}
    ).fillna(0)
    
        # Pondération oedème
    df["lesion_cutane_score"] = df["lesion_cutane"].map(
        {"yes": 3, "no": 0}
    ).fillna(0)

    # Score brut pondéré
    df["health_index_raw"] = (
        df["toux"] * 1 +
        df["fievre"] * 2 +
        df["douleurs_abdominales"] * 2 +
        df["vomissements"] * 2 +
        df["malnutrition_score"] +
        df["diarrhea_score"] +
        df["oedem_score"]+
        df["lesion_cutane_score"]
    )

    # Maximum théorique possible
    max_score = 22
    df["health_index_norm"] = (df["health_index_raw"] / max_score) * 10

    # Catégorisation
    def categorize(score):
        if score <= 3.0:
            return "Faible risque"
        elif score <= 6.0:
            return "Risque modéré"
        else:
            return "Haut risque"

    df["risk_category"] = df["health_index_norm"].apply(categorize)

    return df
#=================================================================================================

def determine_period_status(df, date_col, new_col="period_status", ref_col="date_de_depistage", ref_date="2025-05-01"):
    """
    Détermine le statut de période pour toutes les lignes d'un DataFrame de manière vectorisée.
    
    Args:
        df (pd.DataFrame): DataFrame contenant la colonne de dates
        date_col (str): nom de la colonne contenant les dates
        new_col (str): nom de la nouvelle colonne à créer (défaut: "period_status")
        ref_col (str): nom de la colonne de référence pour comparaison (défaut: "date_de_depistage")
        ref_date (str): date de référence au format YYYY-MM-DD (défaut: "2025-05-01")
        
    Returns:
        pd.DataFrame: DataFrame avec la nouvelle colonne de période ajoutée
    """
    df_result = df.copy()
    
    # Convertir les colonnes de date
    date_series = pd.to_datetime(df[date_col], errors='coerce')
    ref_series = pd.to_datetime(df[ref_col], errors='coerce') if ref_col in df.columns else None
    ref_date_dt = pd.to_datetime(ref_date)
    
    # Extraire le mois de chaque date
    month = date_series.dt.month
    
    # Condition de base pour les périodes
    base_condition = np.where(
        month.between(1, 4),     # Janvier à Avril
        "janv_avril",
        np.where(
            month.between(5, 9), # Mai à Septembre
            "mai_septembre",
            "apres_septembre"    # Octobre à Décembre
        )
    )
    
    # Si ref_col existe, appliquer la logique de comparaison avec ref_date
    if ref_series is not None:
        # Condition: si date < ref_date, utiliser la classification par mois
        # Sinon, utiliser ref_col pour la comparaison
        condition_ref = np.where(
            date_series < ref_date_dt,
            True,  # Utiliser la classification par mois
            date_series < ref_series  # Utiliser ref_col
        )
        
        df_result[new_col] = np.where(
            condition_ref & month.between(1, 4),
            "janv_avril",
            np.where(
                condition_ref & month.between(5, 9),
                "mai_septembre",
                np.where(
                    condition_ref & month.between(10, 12),
                    "apres_septembre",
                    base_condition
                )
            )
        )
    else:
        # Si pas de ref_col, utiliser seulement ref_date
        condition_ref = date_series < ref_date_dt
        df_result[new_col] = np.where(
            condition_ref & month.between(1, 4),
            "janv_avril",
            np.where(
                condition_ref & month.between(5, 9),
                "mai_septembre",
                np.where(
                    condition_ref & month.between(10, 12),
                    "apres_septembre",
                    base_condition
                )
            )
        )
    
    return df_result
#================================================================================================
import pandas as pd

# end_date = "2025-12-31"  # à définir dans ton contexte

def get_age_flag(df: pd.DataFrame, end_date: str, min_months: int = 180, max_months: int = 588) -> pd.DataFrame:
    """
    Ajoute une colonne 'age_flag' au DataFrame selon l'âge en mois à la date de référence.

    Args:
        df (pd.DataFrame): DataFrame contenant la colonne 'date_naissance'
        end_date (str): Date de référence pour le calcul de l'âge (format 'YYYY-MM-DD')
        min_months (int): Âge minimum en mois (par défaut 180 mois = 15 ans)
        max_months (int): Âge maximum en mois (par défaut 588 mois = 49 ans)

    Returns:
        pd.DataFrame: DataFrame original avec une colonne supplémentaire 'age_flag'
                      contenant "Yes" ou "No"
    """
    end_date_dt = pd.to_datetime(end_date)

    def get_flag(dob):
        dob = pd.to_datetime(dob, errors='coerce')
        if pd.isna(dob):
            return "No"
        # Calcul de l'âge en mois
        age_months = (end_date_dt.year - dob.year) * 12 + (end_date_dt.month - dob.month)
        # Si le jour n'est pas encore atteint ce mois-ci, on retire 1 mois
        if end_date_dt.day < dob.day:
            age_months -= 1
        return "yes" if min_months <= age_months <= max_months else "no"

    df["age_flag"] = df["date_naissance"].apply(get_flag)
    return df
#================================================================================================
def _normalize_text_series(s: pd.Series) -> pd.Series:
    s = s.astype(str).str.lower().str.strip().str.replace(r"\s+", " ", regex=True)
    return s.apply(lambda x: unicodedata.normalize("NFKD", x).encode("ascii","ignore").decode("ascii"))

def _similar(a: str, b: str) -> float:
    return 100.0 * SequenceMatcher(None, a, b).ratio()

def detect_duplicates_with_groups(
    df: pd.DataFrame,
    colonnes: List[str],
    threshold: int = 100,
    return_only_duplicates: int = 1,   # 0=uniques, 1=doublons, 2=tous
    keep_most_na: bool = False
) -> pd.DataFrame:
    """
    Détecte des doublons (stricts si threshold=100, sinon fuzzy) sur `colonnes`,
    et retourne un DataFrame avec:
      - duplicate_group_id (int >= 1 pour les groupes; 0 si singleton)
      - duplicate_group_size (taille du groupe)

    Paramètres
    ----------
    df : DataFrame source
    colonnes : colonnes utilisées pour la comparaison
    threshold : 100 => exact (sur texte normalisé) ; <100 => fuzzy
    return_only_duplicates : 
        0 => retourner uniquement les lignes uniques (duplicate_group_id == 0)
        1 => retourner uniquement les doublons (duplicate_group_id > 0)
        2 => retourner toutes les lignes (uniques + doublons)
    keep_most_na : si True, dans chaque groupe de doublons, ne garder qu'un 
                   enregistrement (celui avec le plus de N/A) puis appliquer le mode

    Returns
    -------
    DataFrame enrichi (ou nettoyé si keep_most_na=True)
    """
    if return_only_duplicates not in (0, 1, 2):
        raise ValueError("return_only_duplicates doit être 0, 1 ou 2.")
    if not colonnes:
        raise ValueError("Aucune colonne fournie.")
    manquantes = [c for c in colonnes if c not in df.columns]
    if manquantes:
        raise ValueError(f"Colonnes manquantes dans df: {manquantes}")

    n = len(df)
    if n == 0:
        out = df.copy()
        out["duplicate_group_id"] = pd.Series(dtype=int)
        out["duplicate_group_size"] = pd.Series(dtype=int)
        return out

    # 1) Normalisation des colonnes de comparaison
    df_cmp = df[colonnes].copy()
    for c in colonnes:
        df_cmp[c] = _normalize_text_series(df_cmp[c])

    # 2) Cas strict (exact)
    if threshold >= 100:
        keys = df_cmp.apply(lambda r: tuple(r.values.tolist()), axis=1)
        codes, _ = pd.factorize(keys, sort=True)
        sizes = pd.Series(codes).map(pd.Series(pd.Series(codes).value_counts()))
        group_id = np.where(sizes.values >= 2, codes + 1, 0)
        out = df.copy()
        out["duplicate_group_id"] = group_id
        out["duplicate_group_size"] = np.where(group_id > 0, sizes.values, 1)

        if keep_most_na:
            return _process_keep_most_na(out, return_mode=return_only_duplicates)

        # Filtrage selon le mode
        if return_only_duplicates == 0:
            return out[out["duplicate_group_id"] == 0].reset_index(drop=True)
        elif return_only_duplicates == 1:
            return out[out["duplicate_group_id"] > 0].reset_index(drop=True)
        else:
            return out.reset_index(drop=True)

    # 3) Cas fuzzy : union-find
    parent = list(range(n))
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x
    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    vals = df_cmp.values.astype(str)
    for i in range(n):
        vi = vals[i]
        for j in range(i+1, n):
            vj = vals[j]
            smin = 100.0
            for k in range(len(colonnes)):
                s = _similar(vi[k], vj[k])
                if s < smin:
                    smin = s
                if smin < threshold:
                    break
            if smin >= threshold:
                union(i, j)

    roots = [find(i) for i in range(n)]
    uniq_roots, next_id = {}, 1
    group_id = []
    for r in roots:
        if r not in uniq_roots:
            uniq_roots[r] = next_id
            next_id += 1
        group_id.append(uniq_roots[r])

    counts = pd.Series(group_id).value_counts()
    size_map = counts.to_dict()
    group_id_final = [gid if size_map[gid] >= 2 else 0 for gid in group_id]
    size_final = [size_map[gid] if size_map[gid] >= 2 else 1 for gid in group_id]

    out = df.copy()
    out["duplicate_group_id"] = group_id_final
    out["duplicate_group_size"] = size_final

    if keep_most_na:
        return _process_keep_most_na(out, return_mode=return_only_duplicates)

    if return_only_duplicates == 0:
        return out[out["duplicate_group_id"] == 0].reset_index(drop=True)
    elif return_only_duplicates == 1:
        return out[out["duplicate_group_id"] > 0].reset_index(drop=True)
    else:
        return out.reset_index(drop=True)


def _process_keep_most_na(df_with_groups: pd.DataFrame, return_mode: int = 1) -> pd.DataFrame:
    """
    Conserve, pour chaque groupe de doublons, l'enregistrement avec le plus de N/A.
    Applique ensuite le mode de retour:
      - 0: uniquement les enregistrements uniques (singletons d'origine)
      - 1: uniquement les 'représentants' gardés des groupes de doublons
      - 2: uniques + représentants gardés
    """
    if return_mode not in (0, 1, 2):
        raise ValueError("return_mode doit être 0, 1 ou 2.")

    technical_cols = ['duplicate_group_id', 'duplicate_group_size']
    data_cols = [col for col in df_with_groups.columns if col not in technical_cols]

    total_cols = len(data_cols)
    df_with_groups['na_count'] = df_with_groups[data_cols].isna().sum(axis=1)
    df_with_groups['na_percentage'] = (df_with_groups['na_count'] / total_cols * 100).round(2)

    print(f"📊 Pourcentage de N/A calculé sur {total_cols} colonnes")
    print(f"📈 Statistiques N/A - Min: {df_with_groups['na_percentage'].min()}%, Max: {df_with_groups['na_percentage'].max()}%")

    def keep_most_na_per_group(group):
        if len(group) == 1:
            return group
        group_sorted = group.sort_values(['na_percentage', 'na_count'], ascending=[False, False])
        most_na = group_sorted.iloc[0:1]
        if 'caseid' in group.columns:
            print(f"   Groupe {group['duplicate_group_id'].iloc[0]}: {len(group)} doublons → gardé caseid {most_na['caseid'].iloc[0]} ({most_na['na_percentage'].iloc[0]}% N/A)")
        else:
            print(f"   Groupe {group['duplicate_group_id'].iloc[0]}: {len(group)} doublons → gardé index {most_na.index[0]} ({most_na['na_percentage'].iloc[0]}% N/A)")
        return most_na

    print("🔍 Traitement des groupes de doublons...")
    non_duplicates = df_with_groups[df_with_groups['duplicate_group_id'] == 0].copy()
    duplicates = df_with_groups[df_with_groups['duplicate_group_id'] > 0].copy()

    print(f"   📋 {len(non_duplicates)} enregistrements uniques")
    print(f"   🔍 {len(duplicates)} enregistrements en doublons dans {duplicates['duplicate_group_id'].nunique()} groupes")

    if len(duplicates) > 0:
        kept_duplicates = duplicates.groupby('duplicate_group_id').apply(keep_most_na_per_group).reset_index(drop=True)
    else:
        kept_duplicates = pd.DataFrame(columns=df_with_groups.columns)

    # Réinitialisation des colonnes techniques dans les résultats finaux
    for df_part in (non_duplicates, kept_duplicates):
        if 'na_count' in df_part.columns: df_part.drop(columns=['na_count'], inplace=True, errors='ignore')
        if 'na_percentage' in df_part.columns: df_part.drop(columns=['na_percentage'], inplace=True, errors='ignore')
        df_part['duplicate_group_id'] = 0
        df_part['duplicate_group_size'] = 1

    if return_mode == 0:
        result = non_duplicates.reset_index(drop=True)
    elif return_mode == 1:
        result = kept_duplicates.reset_index(drop=True)
    else:
        result = pd.concat([non_duplicates, kept_duplicates], ignore_index=True)

    print(f"✅ Résultat final: {len(result)} enregistrements")
    print(f"🎯 Tous les duplicate_group_id sont à 0 (uniques ‘après résolution’): {(result['duplicate_group_id'] == 0).all()}")

    if 'caseid' in result.columns:
        caseid_duplicates = result['caseid'].duplicated().sum()
        if caseid_duplicates > 0:
            print(f"⚠️ Attention: {caseid_duplicates} caseid encore en doublon")
            result = result.drop_duplicates(subset=['caseid'], keep='first')
            print(f"🧹 Après nettoyage final: {len(result)} enregistrements")

    return result
#=================================================================================================
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from pathlib import Path

def copy_ranges_by_mapping(
    mapping: dict,
    rows_source_range: str,
    src_file: str = "enrol_table.xlsx",
    dst_file: str = "cluster.xlsx",
    src_sheet: str | None = None,
    dst_sheet: str | None = None,
):
    """
    Copie des données par blocs de cellules d'un fichier Excel source vers un fichier destination.

    Args:
        mapping: dict {"SRC_RANGE": "DST_RANGE", ...}
                 Ex: {"C4:C22": "A4:A22", "B4:B22": "B4:B22", "D4:F22": "C4:C22"}
        rows_source_range: plage servant à calculer le nombre de lignes à copier (ex: "B4:B22")
        src_file: Excel source (par défaut "enrol_table.xlsx")
        dst_file: Excel destination (par défaut "cluster.xlsx")
        src_sheet: nom de feuille source (None = active)
        dst_sheet: nom de feuille destination (None = active)

    Règles:
      - La hauteur copiée = nombre de lignes de rows_source_range (ou nombre de cellules non vides si plus court).
      - Si largeur src == largeur dst -> copie bloc à bloc.
      - Si largeur src > 1 et largeur dst == 1 -> somme horizontale par ligne (collapsing).
      - Sinon -> erreur.
    """
    # Ouvrir classeurs
    wb_src = load_workbook(src_file, data_only=True)
    wb_dst = load_workbook(dst_file)

    ws_src = wb_src[src_sheet] if src_sheet else wb_src.active
    ws_dst = wb_dst[dst_sheet] if dst_sheet else wb_dst.active

    # -------------------------
    # 1) Déterminer n_rows
    # -------------------------
    rs_min_col, rs_min_row, rs_max_col, rs_max_row = range_boundaries(rows_source_range)
    # hauteur théorique
    n_rows = rs_max_row - rs_min_row + 1
    # compter non vides si tu veux tronquer aux non vides en tête
    non_empty = 0
    for r in range(rs_min_row, rs_min_row + n_rows):
        v = ws_src.cell(row=r, column=rs_min_col).value
        if v is not None and (not isinstance(v, str) or v.strip() != ""):
            non_empty += 1
        else:
            # si tu préfères t'arrêter au premier vide, décommente :
            # break
            pass
    # Choix: on garde la hauteur fournie par la plage; si tu veux limiter aux non vides, utilise:
    # n_rows = max(non_empty, 0)

    # -------------------------
    # 2) Boucle de copie
    # -------------------------
    for src_rng, dst_rng in mapping.items():
        smin_c, smin_r, smax_c, smax_r = range_boundaries(src_rng)
        dmin_c, dmin_r, dmax_c, dmax_r = range_boundaries(dst_rng)

        src_width = smax_c - smin_c + 1
        dst_width = dmax_c - dmin_c + 1

        # Vérifs de cohérence hauteur
        dst_height = dmax_r - dmin_r + 1
        if dst_height < n_rows:
            raise ValueError(
                f"Destination {dst_rng} trop courte ({dst_height} lignes) pour copier {n_rows} lignes."
            )

        if src_width == dst_width:
            # Copie 1:1 des blocs
            for rr in range(n_rows):
                for cc in range(src_width):
                    val = ws_src.cell(row=smin_r + rr, column=smin_c + cc).value
                    ws_dst.cell(row=dmin_r + rr, column=dmin_c + cc, value=val)

        elif src_width > 1 and dst_width == 1:
            # Cas collapsing: on somme horizontalement la source pour déposer dans 1 colonne
            for rr in range(n_rows):
                row_sum = 0
                for cc in range(src_width):
                    v = ws_src.cell(row=smin_r + rr, column=smin_c + cc).value
                    if v is None or (isinstance(v, str) and v.strip() == ""):
                        continue
                    try:
                        row_sum += float(v)
                    except Exception:
                        # Si non numérique, on ignore (ou lève une erreur selon tes besoins)
                        pass
                ws_dst.cell(row=dmin_r + rr, column=dmin_c, value=row_sum)
        else:
            raise ValueError(
                f"Incompatibilité largeur: source {src_rng} (w={src_width}) -> destination {dst_rng} (w={dst_width})."
            )

    # Enregistrer
    out_path = Path(dst_file).with_name(Path(dst_file).stem + "_filled.xlsx")
    wb_dst.save(out_path)
    return str(out_path)
#=================================================================================================
def filter_patients(
    df: pd.DataFrame,
    date_threshold: str = "2025-05-01",
    enrolled_value: str = "yes",
    col_is_enrolled: str = "is_enrolled",
    col_nbr_visit_succeed: str = "nbr_visit_succeed",
    col_enrol_date: str = "enrollement_date_de_visite",
    col_admission_date: str = "date_admission",
    coerce_dates: bool = True
) -> pd.DataFrame:
    """
    Filtre les lignes selon la logique:
      (is_enrolled == 'yes' OR nbr_visit_succeed > 0)
      AND
      (enrollement_date_de_visite >= date_threshold OR date_admission >= date_threshold)

    Args:
        df: DataFrame d'entrée.
        date_threshold: Date seuil au format 'YYYY-MM-DD'.
        enrolled_value: Valeur considérée comme 'inscrit'.
        col_*: Noms des colonnes (si différents dans df).
        coerce_dates: Si True, convertit les colonnes de dates en datetime (errors='coerce').

    Returns:
        Un DataFrame filtré.
    """
    # Copie légère pour éviter de modifier df en place
    data = df.copy()

    # Conversion des dates si demandé
    if coerce_dates:
        for c in (col_enrol_date, col_admission_date):
            if c in data.columns:
                data[c] = pd.to_datetime(data[c], errors="coerce")

    # Date seuil en datetime
    dt = pd.to_datetime(date_threshold)

    # Sécurisation des colonnes manquantes
    if col_is_enrolled not in data.columns:
        data[col_is_enrolled] = pd.NA
    if col_nbr_visit_succeed not in data.columns:
        data[col_nbr_visit_succeed] = 0
    if col_enrol_date not in data.columns:
        data[col_enrol_date] = pd.NaT
    if col_admission_date not in data.columns:
        data[col_admission_date] = pd.NaT

    # Convert nbr_visit_succeed to numeric
    if col_nbr_visit_succeed in data.columns:
        data[col_nbr_visit_succeed] = pd.to_numeric(data[col_nbr_visit_succeed], errors='coerce')

    # Condition booléenne
    cond_left = (data[col_is_enrolled] == enrolled_value) | (data[col_nbr_visit_succeed].fillna(0) > 0)
    cond_right = (data[col_enrol_date] >= dt) | (data[col_admission_date] >= dt)
    condition = cond_left & cond_right

    return data[condition]
#=======================================================================================================================

#Function to get the age range

def age_range(age):
    """
    Catégorise l'âge en mois en tranches d'âge pour l'analyse nutritionnelle
    
    Args:
        age (int/float): âge en mois
        
    Returns:
        str: catégorie d'âge
    """
    # Gérer les valeurs NaN ou négatives
    if pd.isna(age) or age < 0:
        return 'Age_non_defini'
    
    if age < 6:
        return 'Enfants<6mois'
    elif 6 <= age <= 23:  # Correction: utiliser <= au lieu de range()
        return 'Enfants_6_23_mois'
    elif 24 <= age <= 66:  # Correction: utiliser <= au lieu de range()
        return 'Enfants_24_59_mois'
    else:
        return 'Enfants_+59_mois'

#OVC['age_range'] = OVC['age'].map(age_range)
#OVC1 = OVC[OVC['age_range'] != '21+']
#====================================================================================================
def combine_columns(df: pd.DataFrame,
                    col1: str,
                    col2: str,
                    col3: str = "col3",
                    na_value=None) -> pd.DataFrame:
    """
    Combine deux colonnes de dates (col1, col2) en une seule (col3)
    en sélectionnant la date la plus récente entre les deux.

    Args:
        df (pd.DataFrame): Le DataFrame d'entrée
        col1 (str): Nom de la première colonne (date)
        col2 (str): Nom de la deuxième colonne (date)
        col3 (str): Nom de la nouvelle colonne à créer
        na_value (str|None): Valeur considérée comme vide ('---' ou None)

    Returns:
        pd.DataFrame: DataFrame avec une nouvelle colonne col3 contenant
                      la date la plus récente entre col1 et col2.
    """
    # 1️⃣ Normalisation des valeurs vides
    if na_value is not None:
        df[col1] = df[col1].replace(na_value, np.nan)
        df[col2] = df[col2].replace(na_value, np.nan)

    # 2️⃣ Conversion en datetime
    df[col1] = pd.to_datetime(df[col1], errors="coerce")
    df[col2] = pd.to_datetime(df[col2], errors="coerce")

    # 3️⃣ Sélection de la date la plus récente (max)
    df[col3] = df[[col1, col2]].max(axis=1)

    return df
#===========================================================================================================
import pandas as pd
import numpy as np

def compute_nut_filtered(
    enroled: pd.DataFrame,
    depistage: pd.DataFrame,
    start_date: str = "2025-05-01",
    end_date: str = "2025-09-30",
    date_format: str = "%Y-%m-%d",
    na_marker=None,
    clamp_dates: bool = True,
    recompute_age: bool = False,       # ← forcer recalcul âge via utils si True
    verbose: bool = True
):
    """
    Filtre les bénéficiaires enrôlés/admis après start_date et construit les colonnes downstream :
      - date_enrollement = max(enrollement_date_de_visite, date_admission)
      - enrrolled_where normalisé
      - age_years / age_months (seulement si recompute_age=True, sinon utilise age_month(s) déjà présent)
      - age_range dérivé d'age_month(s) existant
      - mamba_given en fonction de start_date (yes si date_enrollement >= start_date)

    Retourne:
      nut_filtered (DataFrame), stats (dict)
    """
    df = enroled.copy()

    # -----------------------------
    # 1) Parsing des dates
    # -----------------------------
    for c in ["enrollement_date_de_visite", "date_admission", "date_of_birth"]:
        if c not in df.columns:
            df[c] = pd.NaT
        if na_marker is not None:
            df[c] = df[c].replace(na_marker, np.nan)
        df[c] = pd.to_datetime(df[c], format=date_format, errors="coerce")

    start_ts = pd.to_datetime(start_date)
    end_ts   = pd.to_datetime(end_date)

    # -----------------------------
    # 2) Numérisation compteur visites
    # -----------------------------
    if "nbr_visit_succeed" in df.columns:
        df["nbr_visit_succeed"] = pd.to_numeric(df["nbr_visit_succeed"], errors="coerce").fillna(0)
    else:
        df["nbr_visit_succeed"] = 0

    # -----------------------------
    # 3) Condition d’éligibilité
    # -----------------------------
    is_enrolled = (df.get("is_enrolled", "no") == "yes") | (df["nbr_visit_succeed"] > 0)
    date_ok = (
        (df["enrollement_date_de_visite"] >= start_ts) |
        (df["date_admission"] >= start_ts)
    )
    condition = is_enrolled & date_ok

    nut_filtered = df.loc[condition].copy()

    # -----------------------------
    # 4) date_enrollement = plus récente des deux
    # -----------------------------
    nut_filtered["date_enrollement"] = nut_filtered[
        ["enrollement_date_de_visite", "date_admission"]
    ].max(axis=1)

    if clamp_dates:
        nut_filtered.loc[nut_filtered["date_enrollement"] < start_ts, "date_enrollement"] = start_ts
        nut_filtered.loc[nut_filtered["date_enrollement"] > end_ts,   "date_enrollement"] = end_ts

    # -----------------------------
    # 5) enrrolled_where normalisé
    # -----------------------------
    if "enrrolled_where" in nut_filtered.columns:
        nut_filtered["enrrolled_where"] = (
            nut_filtered["enrrolled_where"]
              .replace("---", "community")
              .fillna("community")
        )
    else:
        nut_filtered["enrrolled_where"] = "community"

    # -----------------------------
    # 6) Âge : utiliser ce qui existe déjà, sinon *optionnellement* recalculer via utils
    # -----------------------------
    # Normaliser le nom de colonne existant (age_month / age_months)
    age_month_col = "age_months" if "age_months" in nut_filtered.columns else (
        "age_month" if "age_month" in nut_filtered.columns else None
    )

    if recompute_age:
        # Recalcul via utils (tes fonctions existantes) – en référence à date_enrollement
        # On suppose que tes utilitaires acceptent (df, col_dob) et calculent vs "today".
        # Si besoin d'une ref date, adapte utils pour accepter une ref.
        nut_filtered = get_age_in_year(nut_filtered, "date_of_birth")
        nut_filtered = get_age_in_months(nut_filtered, "date_of_birth")
        age_month_col = "age_months" if "age_months" in nut_filtered.columns else age_month_col

    if age_month_col is None:
        # Si aucune colonne âge en mois n’existe et recompute_age=False, on crée un placeholder vide
        nut_filtered["age_months"] = pd.Series(pd.NA, index=nut_filtered.index, dtype="Int64")
        age_month_col = "age_months"

    # Classe d'âge dérivée de l'âge en mois existant
    def map_age_range(m):
        if pd.isna(m):
            return np.nan
        if 0 <= m <= 2:
            return "0_2 mois"
        if 2 < m <= 12:
            return "2_12 mois"
        if m > 12:
            return "+12 mois"
        return np.nan

    nut_filtered["age_range"] = nut_filtered[age_month_col].map(map_age_range)

    # -----------------------------
    # 7) Merge dépistage (caseid, username, date)
    # -----------------------------
    deps = depistage.copy()
    deps = deps.rename(columns={
        "form.case.@case_id": "caseid",
        "form.depistage.date_de_visite": "date_de_depistage"
    })
    cols_to_keep = [c for c in ["date_de_depistage", "caseid", "username"] if c in deps.columns]

    if "caseid" in nut_filtered.columns and len(cols_to_keep) > 0:
        nut_filtered = pd.merge(nut_filtered, deps[cols_to_keep], on="caseid", how="left")

    # -----------------------------
    # 8) user_mamba (digits dans username)
    # -----------------------------
    if "username" in nut_filtered.columns:
        nut_filtered["user_mamba"] = nut_filtered["username"].astype(str).str.extract(r"(\d+)")
    else:
        if verbose:
            print("Warning: 'username' column not found")
        nut_filtered["user_mamba"] = pd.NA

    # -----------------------------
    # 9) mamba_given
    #    yes si date_enrollement >= start_date (couvre période et après)
    # -----------------------------
    nut_filtered["mamba_given"] = np.where(
        nut_filtered["date_enrollement"] >= start_ts, "yes", "no"
    )

    stats = {
        "mamba_yes": int((nut_filtered["mamba_given"] == "yes").sum()),
        "mamba_no":  int((nut_filtered["mamba_given"] == "no").sum()),
        "n_rows":    int(len(nut_filtered))
    }

    if verbose:
        print("=== CRÉATION DE LA COLONNE MAMBA_GIVEN ===")
        print(f"Agent avec MAMBA (mamba_given=yes): {stats['mamba_yes']}")
        print(f"Agent non éligibles MAMBA (mamba_given=no): {stats['mamba_no']}")
        print(f"Total lignes retenues: {stats['n_rows']}")

    return nut_filtered, stats

#=================================================================================================
def execute_sql_query(env_path: str, sql_file_path: str) -> pd.DataFrame:
    load_dotenv(env_path)
    user = os.getenv('MYSQL_USER')
    password = os.getenv('MYSQL_PASSWORD')
    host = os.getenv('MYSQL_HOST')
    db = os.getenv('MYSQL_DB')

    conn_text = f'mysql+pymysql://{user}:{password}@{host}/{db}'
    engine = create_engine(conn_text)

    with open(sql_file_path, 'r') as file:
        sql_query = file.read().replace('use caris_db;', '')

    df = pd.read_sql_query(sql_query, engine)
    engine.dispose()

    return df

